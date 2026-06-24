from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Any
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.database.db import get_db
from backend.models.order import Order
from backend.models.supplier import Supplier
from backend.models.material import Material
from backend.models.prediction import Prediction
from backend.services.supplier_service import get_all_suppliers_intelligence
from backend.services.analytics_service import get_analytics_charts_data
from backend.api.dependencies.auth import get_current_user
import datetime

router = APIRouter(prefix="/reports", tags=["Report Generation"])

def create_styled_excel(title: str, headers: list, rows: list) -> io.BytesIO:
    """
    Helper to generate a beautifully styled openpyxl Workbook in memory.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Title Block
    ws.merge_cells("A1:H1")
    ws["A1"] = title.upper()
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark Navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Subtitle with Timestamp
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Confidential Enterprise Report"
    ws["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="333333")
    ws["A2"].fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Empty Row spacer
    ws.row_dimensions[3].height = 10
    
    # 2. Table Headers
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid") # Medium Blue
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 28
    
    # 3. Data Rows
    row_font = Font(name="Segoe UI", size=10, color="2D3748")
    zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row_num, row_data in enumerate(rows, 5):
        fill = zebra_fill if row_num % 2 == 0 else white_fill
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.font = row_font
            cell.fill = fill
            cell.border = thin_border
            
            # Align numbers right, text left
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 22
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        # Skip title and subtitle cells length check
        for cell in col[3:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@router.get("/suppliers")
def export_supplier_report(
    db: Session = Depends(get_db),
    user: Any = Depends(get_current_user)
):
    """
    Exports a comprehensive Supplier Performance and Risk Intelligence Report.
    """
    profiles = get_all_suppliers_intelligence(db)
    
    headers = [
        "Supplier ID", "Supplier Name", "Supplier Tier", "Risk Score (0-100)", 
        "Reliability Score", "Performance Rating", "Total Orders", 
        "Delayed Orders", "Delivery Success Rate", "Average Delay (Days)"
    ]
    
    rows = []
    for p in profiles:
        rows.append([
            p["id"],
            p["name"],
            p["supplier_type"].replace("  ", " - "),
            p["risk_score"],
            p["reliability_score"],
            p["performance_rating"],
            p["total_orders"],
            p["delayed_orders"],
            f"{p['delivery_success_rate']}%",
            p["average_delay_days"]
        ])
        
    buffer = create_styled_excel("Supplier Performance & Risk Report", headers, rows)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Supplier_Performance_Report.xlsx"}
    )

@router.get("/risk-assessment")
def export_risk_report(
    db: Session = Depends(get_db),
    user: Any = Depends(get_current_user)
):
    """
    Exports a platform-wide Risk Assessment Report detailing predicted slippages.
    """
    orders = db.query(Order).all()
    
    headers = [
        "Order ID", "Supplier Name", "Material Name", "Quantity", 
        "Order Date", "Committed Date", "Risk Score (0-100)", "Risk Level", 
        "Delay Probability"
    ]
    
    rows = []
    for o in orders:
        supplier_name = db.query(Supplier.name).filter(Supplier.id == o.supplier_id).scalar() or "Unknown"
        material_name = db.query(Material.material_name).filter(Material.id == o.material_id).scalar() or "Unknown"
        rows.append([
            f"CSC-DEL-000000000{o.id}",
            supplier_name,
            material_name,
            o.quantity,
            o.order_date.isoformat(),
            o.delivery_date.isoformat(),
            o.risk_score,
            o.risk_level,
            f"{int(o.prediction_probability * 100)}%"
        ])
        
    buffer = create_styled_excel("Material Delivery Risk Assessment Report", headers, rows)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Risk_Assessment_Report.xlsx"}
    )

@router.get("/planning")
def export_planning_report(
    db: Session = Depends(get_db),
    user: Any = Depends(get_current_user)
):
    """
    Exports a Procurement Schedule Planning Report.
    """
    orders = db.query(Order).all()
    
    headers = [
        "Order ID", "Material Name", "Category", "Supplier Name", 
        "Quantity", "Required Date", "Lead Days", "Expected Delay", 
        "Safety Buffer", "Recommended Order Date"
    ]
    
    rows = []
    for o in orders:
        supplier = db.query(Supplier).filter(Supplier.id == o.supplier_id).first()
        material = db.query(Material).filter(Material.id == o.material_id).first()
        prediction = db.query(Prediction).filter(Prediction.order_id == o.id).first()
        
        sup_name = supplier.name if supplier else "Unknown"
        mat_name = material.material_name if material else "Unknown"
        cat = material.category if material else "Unknown"
        
        lead_days = 30
        if supplier:
            lead_days = 25 if "Framework" in supplier.supplier_type and "Non-Framework" not in supplier.supplier_type else 30 if "Regional" in supplier.supplier_type else 45
            
        delay_days = prediction.expected_delay_days if prediction else 0
        
        buffer_days = 3
        if o.risk_score > 70:
            buffer_days = 5
            
        total_back = lead_days + delay_days + buffer_days
        rec_order_date = o.delivery_date - datetime.timedelta(days=total_back)
        
        rows.append([
            f"CSC-DEL-000000000{o.id}",
            mat_name,
            cat,
            sup_name,
            o.quantity,
            o.delivery_date.isoformat(),
            lead_days,
            delay_days,
            buffer_days,
            rec_order_date.isoformat()
        ])
        
    buffer = create_styled_excel("Procurement Planning & Scheduling Report", headers, rows)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Procurement_Planning_Report.xlsx"}
    )

@router.get("/project-delays")
def export_project_report(
    db: Session = Depends(get_db),
    user: Any = Depends(get_current_user)
):
    """
    Exports a Project-level Delay Risk and Health Score Report.
    """
    charts_data = get_analytics_charts_data(db)
    heatmap = charts_data["project_risk_heatmap"]
    health = charts_data["project_health_scores"]
    
    headers = [
        "Project/Site Name", "Region Location", "Order Volume", 
        "Risk Index (0-100)", "Project Health Score", "Status Status"
    ]
    
    rows = []
    # Map project names to match them up
    health_dict = {h["project_name"]: h for h in health}
    
    for p in heatmap:
        proj_name = p["project_name"]
        h_info = health_dict.get(proj_name, {"health_score": 100, "status": "EXCELLENT"})
        
        rows.append([
            proj_name,
            p["region"],
            p["order_volume"],
            p["risk_index"],
            h_info["health_score"],
            h_info["status"]
        ])
        
    buffer = create_styled_excel("Project Site Delay Risk & Health Report", headers, rows)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Project_Delay_Report.xlsx"}
    )
